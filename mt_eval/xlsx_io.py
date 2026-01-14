from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from typing import Dict, List, Tuple, Any, Optional
import json
import uuid
from datetime import datetime, timezone

import openpyxl
from openpyxl.workbook.workbook import Workbook

from .config import RunConfig
from .hashing import compute_row_input_hash, stable_seed_int


REQUIRED_SHEETS = ("inputs", "eval")


@dataclass
class WorkbookState:
    wb: Workbook
    inputs_ws_name: str = "inputs"
    eval_ws_name: str = "eval"
    run_id: str = ""
    item_ids: List[int] = None

    # cached headers -> column indices (1-based)
    inputs_col: Dict[str, int] = None
    eval_col: Dict[str, int] = None


def _header_map(ws) -> Dict[str, int]:
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    return {str(name): idx + 1 for idx, name in enumerate(header) if name is not None}


def load_workbook_from_upload(file_bytes: bytes, cfg: RunConfig) -> WorkbookState:
    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    for s in REQUIRED_SHEETS:
        if s not in wb.sheetnames:
            raise ValueError(f"Workbook missing required sheet: '{s}'")

    inputs_ws = wb["inputs"]
    eval_ws = wb["eval"]

    inputs_col = _header_map(inputs_ws)
    eval_col = _header_map(eval_ws)

    # Validate required columns
    _validate_schema(inputs_ws, eval_ws, inputs_col, eval_col, cfg)

    # Validate item_id uniqueness and alignment
    inputs_item_ids = _read_item_ids(inputs_ws, inputs_col["item_id"])
    eval_item_ids = _read_item_ids(eval_ws, eval_col["item_id"])

    if inputs_item_ids != eval_item_ids:
        raise ValueError("item_id mismatch between inputs and eval sheets.")
    if len(inputs_item_ids) != len(set(inputs_item_ids)):
        raise ValueError("Duplicate item_id detected.")

    run_id = _get_or_init_run_id(eval_ws, eval_col, inputs_item_ids)

    state = WorkbookState(
        wb=wb,
        run_id=run_id,
        item_ids=inputs_item_ids,
        inputs_col=inputs_col,
        eval_col=eval_col,
    )

    # Hard-fail hash validation (inputs hash recompute + cross-sheet match)
    _hard_validate_input_hashes(state, cfg)

    # Ensure display_map_json exists per row (deterministic per sentence)
    _ensure_display_maps(state, cfg)

    return state


def save_workbook_to_bytes(state: WorkbookState) -> bytes:
    bio = BytesIO()
    state.wb.save(bio)
    return bio.getvalue()


def now_iso_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


# ---------------- Schema validation ----------------

def _validate_schema(inputs_ws, eval_ws, ic: Dict[str, int], ec: Dict[str, int], cfg: RunConfig) -> None:
    required_inputs = ["item_id", "source", "row_input_hash"] + [f"t{i}" for i in range(1, cfg.num_translations + 1)]
    for col in required_inputs:
        if col not in ic:
            raise ValueError(f"inputs sheet missing required column: {col}")

    required_eval = (
        ["item_id", "comment", "started_at", "committed_at", "edit_count", "display_map_json", "row_input_hash", "row_eval_hash", "run_id"]
        + [f"bucket_t{i}" for i in range(1, cfg.num_translations + 1)]
        + [f"da_t{i}" for i in range(1, cfg.num_translations + 1)]
    )
    for col in required_eval:
        if col not in ec:
            raise ValueError(f"eval sheet missing required column: {col}")


def _read_item_ids(ws, item_id_col: int) -> List[int]:
    ids: List[int] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        v = row[item_id_col - 1]
        if v is None:
            continue
        try:
            ids.append(int(v))
        except Exception:
            raise ValueError(f"Non-integer item_id: {v}")
    return ids


def _get_or_init_run_id(eval_ws, ec: Dict[str, int], item_ids: List[int]) -> str:
    run_col = ec["run_id"]
    run_ids = set()
    for r_idx in range(2, 2 + len(item_ids)):
        v = eval_ws.cell(row=r_idx, column=run_col).value
        if v:
            run_ids.add(str(v))

    if len(run_ids) > 1:
        raise ValueError("Multiple run_id values found in eval sheet (expected 0 or 1).")

    if len(run_ids) == 1:
        return next(iter(run_ids))

    new_id = str(uuid.uuid4())
    for r_idx in range(2, 2 + len(item_ids)):
        eval_ws.cell(row=r_idx, column=run_col).value = new_id
    return new_id


# ---------------- Hash validation ----------------

def _hard_validate_input_hashes(state: WorkbookState, cfg: RunConfig) -> None:
    inputs_ws = state.wb[state.inputs_ws_name]
    eval_ws = state.wb[state.eval_ws_name]
    ic = state.inputs_col
    ec = state.eval_col

    for i, item_id in enumerate(state.item_ids):
        r = 2 + i

        source = str(inputs_ws.cell(row=r, column=ic["source"]).value or "")
        translations = []
        for k in range(1, cfg.num_translations + 1):
            translations.append(str(inputs_ws.cell(row=r, column=ic[f"t{k}"]).value or ""))

        recomputed = compute_row_input_hash(source, translations)

        inputs_hash = str(inputs_ws.cell(row=r, column=ic["row_input_hash"]).value or "")
        eval_hash = str(eval_ws.cell(row=r, column=ec["row_input_hash"]).value or "")

        if not inputs_hash:
            raise ValueError(f"Missing inputs.row_input_hash at item_id={item_id}")
        if inputs_hash != recomputed:
            raise ValueError(f"Hash mismatch (inputs recompute) at item_id={item_id}")
        if eval_hash and eval_hash != inputs_hash:
            raise ValueError(f"Hash mismatch (inputs vs eval) at item_id={item_id}")

        # Copy into eval if blank
        if not eval_hash:
            eval_ws.cell(row=r, column=ec["row_input_hash"]).value = inputs_hash


# ---------------- Display map ----------------

def _ensure_display_maps(state: WorkbookState, cfg: RunConfig) -> None:
    eval_ws = state.wb[state.eval_ws_name]
    ec = state.eval_col

    for i, item_id in enumerate(state.item_ids):
        r = 2 + i
        v = eval_ws.cell(row=r, column=ec["display_map_json"]).value
        if v:
            # Must be valid JSON mapping "1".."N" -> "t#"
            try:
                m = json.loads(v)
                if not isinstance(m, dict) or len(m) != cfg.num_translations:
                    raise ValueError
            except Exception:
                raise ValueError(f"Malformed display_map_json at item_id={item_id}")
            continue

        # Create deterministic permutation of t1..tN
        t_cols = [f"t{k}" for k in range(1, cfg.num_translations + 1)]
        seed = stable_seed_int(state.run_id, int(item_id))
        import random
        rng = random.Random(seed)
        rng.shuffle(t_cols)

        m = {str(pos): t_cols[pos - 1] for pos in range(1, cfg.num_translations + 1)}
        eval_ws.cell(row=r, column=ec["display_map_json"]).value = json.dumps(m, ensure_ascii=False, separators=(",", ":"))
