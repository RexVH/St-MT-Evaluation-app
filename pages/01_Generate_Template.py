from __future__ import annotations

import csv
import io
import json
import random
import re
import uuid
from datetime import datetime, timezone
from typing import Dict, List, Optional, Tuple

import openpyxl
import streamlit as st
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from mt_eval.config import RunConfig, load_config
from mt_eval.hashing import compute_row_input_hash


# -----------------------------
# Utilities
# -----------------------------
def now_iso_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def latin_filler(seed: int, min_words: int = 8, max_words: int = 18) -> str:
    """Deterministic pseudo-latin filler text (placeholder)."""
    rng = random.Random(seed)
    vocab = [
        "lorem", "ipsum", "dolor", "sit", "amet", "consectetur",
        "adipiscing", "elit", "sed", "do", "eiusmod", "tempor",
        "incididunt", "ut", "labore", "et", "dolore", "magna",
        "aliqua", "enim", "ad", "minim", "veniam", "quis",
        "nostrud", "exercitation", "ullamco", "laboris",
        "nisi", "aliquip", "ex", "ea", "commodo", "consequat",
        "duis", "aute", "irure", "in", "reprehenderit",
        "voluptate", "velit", "esse", "cillum", "eu", "fugiat",
        "nulla", "pariatur",
    ]
    n = rng.randint(min_words, max_words)
    words = [rng.choice(vocab) for _ in range(n)]
    s = " ".join(words)
    return s[:1].upper() + s[1:] + "."


def normalize_rows_to_n_translations(
    sources: List[str],
    t_cols: Dict[int, List[str]],
    n: int,
) -> List[Tuple[int, str, List[str]]]:
    """Return list of tuples: (item_id, source, [t1..tN]) in given order."""
    rows: List[Tuple[int, str, List[str]]] = []
    for i, src in enumerate(sources, start=1):
        ts: List[str] = []
        for k in range(1, n + 1):
            col = t_cols.get(k, [])
            ts.append(col[i - 1] if i - 1 < len(col) else "")
        rows.append((i, src, ts))
    return rows


# -----------------------------
# Workbook generation
# -----------------------------
def _eval_header(cfg: RunConfig) -> List[str]:
    """
    Eval sheet schema expected by mt_eval.xlsx_io / app.py:

    item_id,
    bucket_t1..bucket_tN,
    da_t1..da_tN,
    comment, started_at, committed_at, edit_count,
    display_map_json, row_input_hash, row_eval_hash, run_id
    """
    n = int(cfg.num_translations)
    return (
        ["item_id"]
        + [f"bucket_t{i}" for i in range(1, n + 1)]
        + [f"da_t{i}" for i in range(1, n + 1)]
        + [
            "comment",
            "started_at",
            "committed_at",
            "edit_count",
            "display_map_json",
            "row_input_hash",
            "row_eval_hash",
            "run_id",
        ]
    )


def build_workbook(
    rows: List[Tuple[int, str, List[str]]],
    cfg: RunConfig,
    run_id: Optional[str] = None,
) -> Workbook:
    if run_id is None:
        run_id = str(uuid.uuid4())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    inputs = wb.create_sheet("inputs")
    evals = wb.create_sheet("eval")

    n = int(cfg.num_translations)

    inputs_header = ["item_id", "source"] + [f"t{k}" for k in range(1, n + 1)] + ["row_input_hash"]
    inputs.append(inputs_header)

    evals.append(_eval_header(cfg))

    for (item_id, source, translations) in rows:
        translations = (translations + [""] * n)[:n]

        # IMPORTANT: compute_row_input_hash signature is (source: str, translations: List[str])
        row_input_hash = compute_row_input_hash(str(source), [str(t) for t in translations])

        inputs.append([item_id, source, *translations, row_input_hash])

        # defaults: bucket/da empty; other fields blank; edit_count=0; row_eval_hash blank; run_id set
        evals.append(
            [item_id]
            + [""] * n
            + [""] * n
            + ["", "", "", 0, "", row_input_hash, "", run_id]
        )

    inputs.freeze_panes = "A2"
    evals.freeze_panes = "A2"

    # light formatting so it's readable
    inputs.column_dimensions["A"].width = 10
    inputs.column_dimensions["B"].width = 60
    for k in range(1, n + 1):
        col_letter = get_column_letter(2 + k)
        inputs.column_dimensions[col_letter].width = 55
    inputs.column_dimensions[get_column_letter(2 + n + 1)].width = 72

    evals.column_dimensions["A"].width = 10
    for idx in range(2, 2 + (2 * n)):
        evals.column_dimensions[get_column_letter(idx)].width = 12
    for idx in range(2 + 2 * n, 2 + 2 * n + 9):
        evals.column_dimensions[get_column_letter(idx)].width = 30

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# -----------------------------
# Upload table parsing (CSV/TSV/XLSX) -> rows
# -----------------------------
_TCOL_RE = re.compile(r"^t(\d+)$", re.IGNORECASE)


def _canon(s: object) -> str:
    return str(s).strip() if s is not None else ""


def _detect_required_columns(headers: List[str], cfg: RunConfig) -> Tuple[int, Dict[int, int], List[str]]:
    """
    Returns:
      src_idx: index in headers for source/src (0-based)
      t_idx_map: {k -> index in headers} for t1..tN
      normalized_headers: for debug display
    Raises on missing/extra t columns.
    """
    n = int(cfg.num_translations)
    norm = [_canon(h) for h in headers]
    norm_low = [h.lower() for h in norm]

    # source/src (case-insensitive)
    src_idx = -1
    for i, h in enumerate(norm_low):
        if h == "source" or h == "src":
            src_idx = i
            break
    if src_idx < 0:
        raise ValueError("Missing required column: 'source' (case-insensitive; 'src' also accepted).")

    # translation columns: strict t1..tN must exist; no t{>N}
    t_idx_map: Dict[int, int] = {}
    found_t_nums: List[int] = []
    for i, h in enumerate(norm_low):
        m = _TCOL_RE.match(h)
        if not m:
            continue
        k = int(m.group(1))
        found_t_nums.append(k)
        if k <= n:
            t_idx_map[k] = i

    missing = [k for k in range(1, n + 1) if k not in t_idx_map]
    extra = sorted([k for k in found_t_nums if k > n])
    if missing:
        raise ValueError(f"Missing translation columns: {', '.join([f't{k}' for k in missing])} (expected exactly t1..t{n}).")
    if extra:
        raise ValueError(f"Found unexpected translation columns beyond config: {', '.join([f't{k}' for k in extra])} (config expects t1..t{n}).")

    return src_idx, t_idx_map, norm


def _row_is_all_blank(src: str, translations: List[str]) -> bool:
    if src.strip() != "":
        return False
    return all(t.strip() == "" for t in translations)


def _extract_and_validate_rows(
    headers: List[str],
    raw_rows: List[List[object]],
    cfg: RunConfig,
) -> Tuple[List[Tuple[int, str, List[str]]], List[str]]:
    src_idx, t_idx_map, norm_headers = _detect_required_columns(headers, cfg)
    n = int(cfg.num_translations)

    rows: List[Tuple[int, str, List[str]]] = []
    for ridx, raw in enumerate(raw_rows, start=2):  # for error messages: header is row 1
        if len(raw) < len(headers):
            raw = list(raw) + [""] * (len(headers) - len(raw))

        src = _canon(raw[src_idx]).strip()
        translations = [_canon(raw[t_idx_map[k]]).strip() for k in range(1, n + 1)]

        # Stop on the FIRST fully-blank row
        if _row_is_all_blank(src, translations):
            break

        any_t = any(t.strip() != "" for t in translations)
        all_t_blank = not any_t

        # Error: source present but all translations blank
        if src.strip() != "" and all_t_blank:
            raise ValueError(
                f"Row {ridx}: source/src is present but ALL translations t1..t{n} are blank."
            )

        # Error: source blank but at least one translation present
        if src.strip() == "" and any_t:
            raise ValueError(
                f"Row {ridx}: one or more translations are present but source/src is blank."
            )

        item_id = len(rows) + 1
        rows.append((item_id, src, translations))

    if not rows:
        raise ValueError("No data rows found (or the first data row was blank).")

    return rows, norm_headers


def parse_csv_or_tsv(upload_bytes: bytes, cfg: RunConfig) -> Tuple[List[Tuple[int, str, List[str]]], List[str]]:
    text = upload_bytes.decode("utf-8-sig", errors="replace")
    sio = io.StringIO(text)

    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"])
    except Exception:
        dialect = csv.excel

    reader = csv.reader(sio, dialect)
    try:
        headers = next(reader)
    except StopIteration:
        raise ValueError("File appears to be empty (no header row).")

    raw_rows: List[List[object]] = list(reader)
    return _extract_and_validate_rows(headers, raw_rows, cfg)


def parse_xlsx_table(upload_bytes: bytes, cfg: RunConfig) -> Tuple[List[Tuple[int, str, List[str]]], List[str]]:
    # data_only=True reads cached computed values for formulas (when available)
    wb = openpyxl.load_workbook(io.BytesIO(upload_bytes), data_only=True)
    ws = wb.worksheets[0]

    headers: List[str] = [(cell.value if cell.value is not None else "") for cell in ws[1]]

    raw_rows: List[List[object]] = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        raw_rows.append(list(row))

    return _extract_and_validate_rows(headers, raw_rows, cfg)


def parse_uploaded_table(uploaded_name: str, upload_bytes: bytes, cfg: RunConfig) -> Tuple[List[Tuple[int, str, List[str]]], List[str]]:
    name = (uploaded_name or "").lower()
    if name.endswith(".xlsx"):
        return parse_xlsx_table(upload_bytes, cfg)
    if name.endswith(".csv") or name.endswith(".tsv") or name.endswith(".txt"):
        return parse_csv_or_tsv(upload_bytes, cfg)

    # fallback: try CSV then XLSX
    try:
        return parse_csv_or_tsv(upload_bytes, cfg)
    except Exception:
        return parse_xlsx_table(upload_bytes, cfg)


# -----------------------------
# XLSX Repair/Rebuild helpers
# -----------------------------
def header_map(ws) -> Dict[str, int]:
    header = [cell.value for cell in ws[1]]
    out: Dict[str, int] = {}
    for idx, name in enumerate(header, start=1):
        if name is None:
            continue
        out[str(name).strip()] = idx
    return out


def ensure_column(ws, name: str) -> int:
    """Ensure a header column exists; if missing, append it. Returns 1-based col index."""
    hm = header_map(ws)
    if name in hm:
        return hm[name]
    ws.cell(row=1, column=ws.max_column + 1, value=name)
    return ws.max_column


def read_inputs_rows(wb: Workbook, cfg: RunConfig) -> List[Dict[str, object]]:
    if "inputs" not in wb.sheetnames:
        raise ValueError("Workbook missing required sheet: 'inputs'")

    ws = wb["inputs"]
    hm = header_map(ws)

    required = ["item_id", "source"] + [f"t{k}" for k in range(1, int(cfg.num_translations) + 1)]
    missing = [c for c in required if c not in hm]
    if missing:
        raise ValueError(f"inputs sheet missing columns: {missing}")

    rows: List[Dict[str, object]] = []
    for r in range(2, ws.max_row + 1):
        item_id = ws.cell(r, hm["item_id"]).value
        if item_id is None or str(item_id).strip() == "":
            continue
        row = {"item_id": int(item_id), "source": ws.cell(r, hm["source"]).value or ""}
        for k in range(1, int(cfg.num_translations) + 1):
            row[f"t{k}"] = ws.cell(r, hm[f"t{k}"]).value or ""
        rows.append(row)

    rows.sort(key=lambda d: int(d["item_id"]))
    return rows


def write_inputs_hashes(wb: Workbook, cfg: RunConfig, rows: List[Dict[str, object]]) -> None:
    ws = wb["inputs"]
    hm = header_map(ws)
    col_hash = ensure_column(ws, "row_input_hash")

    rownum_by_item: Dict[int, int] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, hm["item_id"]).value
        if v is None or str(v).strip() == "":
            continue
        rownum_by_item[int(v)] = r

    n = int(cfg.num_translations)
    for row in rows:
        source = str(row.get("source") or "")
        translations = [str(row.get(f"t{k}") or "") for k in range(1, n + 1)]
        rh = compute_row_input_hash(source, translations)
        rnum = rownum_by_item.get(int(row["item_id"]))
        if rnum:
            ws.cell(row=rnum, column=col_hash, value=rh)


def rebuild_eval_sheet(wb: Workbook, cfg: RunConfig, rows: List[Tuple[int, str, List[str]]], run_id: str) -> None:
    if "eval" in wb.sheetnames:
        wb.remove(wb["eval"])
    ws = wb.create_sheet("eval")

    ws.append(_eval_header(cfg))
    n = int(cfg.num_translations)

    for (item_id, source, translations) in rows:
        translations = (translations + [""] * n)[:n]
        row_input_hash = compute_row_input_hash(str(source), [str(t) for t in translations])
        ws.append(
            [item_id]
            + [""] * n
            + [""] * n
            + ["", "", "", 0, "", row_input_hash, "", run_id]
        )

    ws.freeze_panes = "A2"


def repair_eval_sheet_in_place(wb: Workbook, cfg: RunConfig, rows: List[Tuple[int, str, List[str]]], run_id: str) -> None:
    """Keep existing eval data where possible; ensure required columns/rows; repair row_input_hash/run_id."""
    if "eval" not in wb.sheetnames:
        rebuild_eval_sheet(wb, cfg, rows, run_id)
        return

    ws = wb["eval"]

    for col in _eval_header(cfg):
        ensure_column(ws, col)

    hm = header_map(ws)
    n = int(cfg.num_translations)

    eval_row_by_item: Dict[int, int] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=hm["item_id"]).value
        if v is None or str(v).strip() == "":
            continue
        try:
            eval_row_by_item[int(v)] = r
        except Exception:
            continue

    for (item_id, source, translations) in rows:
        translations = (translations + [""] * n)[:n]
        row_input_hash = compute_row_input_hash(str(source), [str(t) for t in translations])

        rr = eval_row_by_item.get(item_id)
        if rr is None:
            ws.append(
                [item_id]
                + [""] * n
                + [""] * n
                + ["", "", "", 0, "", row_input_hash, "", run_id]
            )
            rr = ws.max_row
            eval_row_by_item[item_id] = rr

        ws.cell(row=rr, column=hm["row_input_hash"]).value = row_input_hash
        ws.cell(row=rr, column=hm["run_id"]).value = run_id

        if ws.cell(row=rr, column=hm["edit_count"]).value in (None, ""):
            ws.cell(row=rr, column=hm["edit_count"]).value = 0

        dm = ws.cell(row=rr, column=hm["display_map_json"]).value
        if dm:
            try:
                m = json.loads(str(dm))
                if not isinstance(m, dict):
                    ws.cell(row=rr, column=hm["display_map_json"]).value = ""
            except Exception:
                ws.cell(row=rr, column=hm["display_map_json"]).value = ""


# -----------------------------
# App
# -----------------------------
def app():
    st.set_page_config(page_title="MT Eval – Template Tools", layout="wide")
    st.title("MT Evaluation – Template Tools (01)")

    cfg: RunConfig = load_config("config.yaml")
    st.caption(f"Configured for {cfg.num_translations} translations per sentence (t1..t{cfg.num_translations}).")

    with st.expander("Current config (debug)", expanded=False):
        # IMPORTANT: only pass JSON-serializable structures (NO cfg object, NO Bucket objects)
        st.json(
            {
                "num_translations": int(cfg.num_translations),
                "bucket_keys": list(cfg.bucket_keys),
                "buckets": [{"key": b.key, "label": b.label} for b in cfg.buckets],
                "validation": {
                    "enforce_bucket_ordering": bool(getattr(cfg, "enforce_bucket_ordering", False)),
                    "allow_empty_buckets": bool(getattr(cfg, "allow_empty_buckets", False)),
                },
                "da_intra_bucket_options": int(getattr(cfg, "da_intra_bucket_options", 3) or 3),
                "bucket_colors": getattr(cfg, "bucket_colors", None),
            },
            expanded=False,
        )

    tab1, tab2, tab3 = st.tabs(
        ["Generate example (5 rows)", "Upload (CSV/XLSX) → XLSX", "Upload XLSX → Repair/Rebuild"]
    )

    with tab1:
        st.subheader("Example file generator")
        st.write("Creates 5 example source sentences and placeholder translations per sentence.")

        if st.button("Generate example XLSX", type="primary"):
            sources = [
                "The quick brown fox jumps over the lazy dog.",
                "I would like to book a table for two at 7 PM.",
                "Please translate this sentence as naturally as possible.",
                "The model produced multiple candidates with varying fluency.",
                "Human evaluation should be consistent and repeatable.",
            ]

            t_cols: Dict[int, List[str]] = {}
            for k in range(1, int(cfg.num_translations) + 1):
                t_cols[k] = [latin_filler(seed=10_000 * k + i) for i in range(1, 6)]

            rows = normalize_rows_to_n_translations(sources, t_cols, int(cfg.num_translations))
            wb = build_workbook(rows, cfg)
            data = workbook_to_bytes(wb)

            st.download_button(
                "Download example XLSX",
                data=data,
                file_name="mt_eval_example_5rows.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    with tab2:
        st.subheader("Upload (CSV/XLSX) → XLSX template")
        st.write(
            "Upload a CSV/TSV or XLSX containing a `source` column (case-insensitive; `src` also accepted) and "
            f"translation columns `t1..t{cfg.num_translations}`.\n\n"
            "- Row order is preserved exactly as provided.\n"
            "- Import stops at the first fully-blank row.\n"
            "- If source is present but **all** translations are blank (or vice versa), the import errors.\n"
            "- Individual missing translations are allowed (kept as blank) as long as at least one translation is present."
        )

        up = st.file_uploader(
            "Upload file",
            type=["csv", "tsv", "txt", "xlsx"],
            accept_multiple_files=False,
            key="table_uploader",
        )

        colA, colB = st.columns([1, 1])
        with colA:
            run_id_override = st.text_input("Optional run_id (leave blank to auto-generate)", value="", key="table_run_id")
        with colB:
            out_name = st.text_input("Output filename", value="mt_eval_from_upload.xlsx", key="table_out_name")

        debug_box = st.empty()

        if up is not None:
            try:
                upload_bytes = up.read()
                rows, detected_cols = parse_uploaded_table(up.name, upload_bytes, cfg)

                st.success(f"Loaded {len(rows)} row(s).")
                st.write("Detected columns (header row):")
                st.code(", ".join([c for c in detected_cols if c.strip()][:200]))

                st.write("Preview (first 5 rows):")
                preview = []
                for (item_id, src, ts) in rows[:5]:
                    row_preview = {"item_id": item_id, "source": src}
                    if ts:
                        row_preview["t1"] = ts[0]
                        row_preview[f"t{cfg.num_translations}"] = ts[int(cfg.num_translations) - 1]
                        row_preview["nonempty_translations"] = sum(1 for x in ts if str(x).strip() != "")
                    preview.append(row_preview)
                st.dataframe(preview, width="stretch")

                # Debug prints (server logs)
                print(f"[DEBUG] Upload→XLSX: name={up.name} bytes={len(upload_bytes)} rows={len(rows)} n={cfg.num_translations}")
                if rows:
                    i0, s0, t0 = rows[0]
                    print(f"[DEBUG] First row: item_id={i0} source_len={len(s0)} nonempty_t={sum(1 for x in t0 if str(x).strip())}")

                wb = build_workbook(rows, cfg, run_id=run_id_override.strip() or None)
                data = workbook_to_bytes(wb)

                st.download_button(
                    "Download XLSX template",
                    data=data,
                    file_name=out_name if out_name.lower().endswith(".xlsx") else out_name + ".xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                print("[DEBUG] Exception during Upload→XLSX:", repr(e))
                debug_box.error(f"Failed to generate XLSX: {e}")

    with tab3:
        st.subheader("Upload XLSX → Repair/Rebuild")
        st.write(
            "Upload an XLSX that contains an `inputs` sheet. This tool will recompute `row_input_hash` for each row. "
            "You can also rebuild or repair the `eval` sheet so the main rating app can load it reliably."
        )

        xlsx_up = st.file_uploader("Upload XLSX", type=["xlsx"], accept_multiple_files=False, key="xlsx_uploader")

        c1, c2, c3 = st.columns([1, 1, 1.2])
        with c1:
            rebuild_eval = st.checkbox("Rebuild eval sheet from scratch", value=False)
        with c2:
            run_id_override_xlsx = st.text_input("Optional run_id override", value="", key="xlsx_run_id")
        with c3:
            out_xlsx_name = st.text_input("Output filename", value="mt_eval_repaired.xlsx", key="xlsx_out_name")

        if xlsx_up is not None:
            try:
                wb = openpyxl.load_workbook(io.BytesIO(xlsx_up.read()))
                rows_dict = read_inputs_rows(wb, cfg)

                # Determine run_id
                run_id = run_id_override_xlsx.strip()
                if not run_id:
                    run_id = str(uuid.uuid4())
                    if "eval" in wb.sheetnames:
                        ev = wb["eval"]
                        hm = header_map(ev)
                        if "run_id" in hm:
                            for r in range(2, ev.max_row + 1):
                                v = ev.cell(row=r, column=hm["run_id"]).value
                                if v and str(v).strip():
                                    run_id = str(v).strip()
                                    break

                write_inputs_hashes(wb, cfg, rows_dict)

                rows_tuple = [
                    (
                        int(r["item_id"]),
                        str(r["source"]),
                        [str(r.get(f"t{k}") or "") for k in range(1, int(cfg.num_translations) + 1)],
                    )
                    for r in rows_dict
                ]
                if rebuild_eval:
                    rebuild_eval_sheet(wb, cfg, rows_tuple, run_id)
                else:
                    repair_eval_sheet_in_place(wb, cfg, rows_tuple, run_id)

                data = workbook_to_bytes(wb)

                st.success(
                    f"Processed {len(rows_dict)} input row(s). "
                    f"{'Rebuilt' if rebuild_eval else 'Repaired'} eval sheet. "
                    f"run_id={run_id}"
                )

                st.download_button(
                    "Download repaired XLSX",
                    data=data,
                    file_name=out_xlsx_name if out_xlsx_name.lower().endswith(".xlsx") else out_xlsx_name + ".xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            except Exception as e:
                st.error(f"Failed to repair/rebuild XLSX: {e}")


if __name__ == "__main__":
    app()
