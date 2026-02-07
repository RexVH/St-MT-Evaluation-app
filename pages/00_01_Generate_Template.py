from __future__ import annotations

import io
import uuid
import json
import random
from datetime import datetime, timezone
from typing import List, Dict, Tuple, Optional

import streamlit as st
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from mt_eval.config import load_config, RunConfig
from mt_eval.hashing import compute_row_input_hash


# -----------------------------
# Utilities
# -----------------------------

def now_iso_utc() -> str:
    return datetime.now(timezone.utc).isoformat()


def latin_filler(seed: int, min_words: int = 8, max_words: int = 18) -> str:
    """
    Deterministic pseudo-latin filler text (NOT true Latin, just placeholder).
    """
    rng = random.Random(seed)
    vocab = [
        "lorem", "ipsum", "dolor", "sit", "amet", "consectetur",
        "adipiscing", "elit", "sed", "do", "eiusmod", "tempor",
        "incididunt", "ut", "labore", "et", "dolore", "magna",
        "aliqua", "enim", "ad", "minim", "veniam", "quis",
        "nostrud", "exercitation", "ullamco", "laboris",
        "nisi", "aliquip", "ex", "ea", "commodo", "consequat",
        "duis", "aute", "irure", "in", "reprehenderit",
        "voluptate", "velit", "esse", "cillum", "eu", "fugiat",
        "nulla", "pariatur",
    ]
    n = rng.randint(min_words, max_words)
    words = [rng.choice(vocab) for _ in range(n)]
    s = " ".join(words)
    return s[:1].upper() + s[1:] + "."


def normalize_rows_to_n_translations(
    sources: List[str],
    t_cols: Dict[int, List[str]],
    n: int,
) -> List[Tuple[int, str, List[str]]]:
    """
    Returns list of tuples: (item_id, source, [t1..tN])
    Missing translations become "".
    Extra translations beyond N are ignored.
    """
    rows: List[Tuple[int, str, List[str]]] = []
    for i, src in enumerate(sources, start=1):
        ts: List[str] = []
        for k in range(1, n + 1):
            col = t_cols.get(k, [])
            ts.append(col[i - 1] if i - 1 < len(col) else "")
        rows.append((i, src, ts))
    return rows


def _eval_header(cfg: RunConfig) -> List[str]:
    n = cfg.num_translations
    return (
        ["item_id"]
        + [f"bucket_t{k}" for k in range(1, n + 1)]
        + [f"da_t{k}" for k in range(1, n + 1)]
        + [
            "comment",
            "started_at",
            "committed_at",
            "edit_count",
            "display_map_json",
            "row_input_hash",
            "row_eval_hash",
            "run_id",
        ]
    )


def build_workbook(rows: List[Tuple[int, str, List[str]]], cfg: RunConfig, run_id: Optional[str] = None) -> Workbook:
    """
    Builds the XLSX with required sheets/columns.
    - inputs: item_id, source, t1..tN, row_input_hash
    - eval: item_id, bucket_t#, da_t#, comment, started_at, committed_at, edit_count,
            display_map_json, row_input_hash, row_eval_hash, run_id
    """
    if run_id is None:
        run_id = str(uuid.uuid4())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    inputs = wb.create_sheet("inputs")
    evals = wb.create_sheet("eval")

    n = cfg.num_translations

    inputs_header = ["item_id", "source"] + [f"t{k}" for k in range(1, n + 1)] + ["row_input_hash"]
    inputs.append(inputs_header)

    evals.append(_eval_header(cfg))

    for (item_id, source, translations) in rows:
        translations = (translations + [""] * n)[:n]
        row_input_hash = compute_row_input_hash(source, translations)

        inputs.append([item_id, source, *translations, row_input_hash])

        evals.append(
            [item_id]
            + [""] * n
            + [""] * n
            + ["", "", "", 0, "", row_input_hash, "", run_id]
        )

    inputs.freeze_panes = "A2"
    evals.freeze_panes = "A2"

    inputs.column_dimensions["A"].width = 10
    inputs.column_dimensions["B"].width = 60
    for k in range(1, n + 1):
        col_letter = get_column_letter(2 + k)
        inputs.column_dimensions[col_letter].width = 55
    inputs.column_dimensions[get_column_letter(2 + n + 1)].width = 72

    evals.column_dimensions["A"].width = 10
    for idx in range(2, 2 + (2 * n)):
        evals.column_dimensions[get_column_letter(idx)].width = 12
    for idx in range(2 + 2 * n, 2 + 2 * n + 9):
        evals.column_dimensions[get_column_letter(idx)].width = 30

    return wb


def workbook_to_bytes(wb: Workbook) -> bytes:
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def parse_csv(upload_bytes: bytes, cfg: RunConfig) -> List[Tuple[int, str, List[str]]]:
    """
    Accepts CSV with:
      - required: source
      - translations: either t1..tN (preferred) or any subset; missing filled with "".
    """
    import csv

    text = upload_bytes.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))

    if not reader.fieldnames:
        raise ValueError("CSV appears to have no header row.")

    fieldnames = [f.strip() for f in reader.fieldnames if f is not None]
    if "source" not in fieldnames:
        raise ValueError("CSV must include a 'source' column.")

    sources: List[str] = []
    t_cols: Dict[int, List[str]] = {}

    for row in reader:
        src = (row.get("source") or "").strip()
        sources.append(src)

        for k in range(1, cfg.num_translations + 1):
            key = f"t{k}"
            if key in row:
                t_cols.setdefault(k, []).append((row.get(key) or "").strip())

    if not sources:
        raise ValueError("CSV has no data rows.")

    return normalize_rows_to_n_translations(sources, t_cols, cfg.num_translations)


# -----------------------------
# XLSX Repair/Rebuild helpers
# -----------------------------

def header_map(ws) -> Dict[str, int]:
    header = []
    for cell in ws[1]:
        header.append(cell.value)
    out: Dict[str, int] = {}
    for idx, name in enumerate(header, start=1):
        if name is None:
            continue
        out[str(name).strip()] = idx
    return out


def ensure_column(ws, name: str) -> int:
    """
    Ensure a header column exists; if missing, append it.
    Returns 1-based column index.
    """
    hm = header_map(ws)
    if name in hm:
        return hm[name]
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col).value = name
    return new_col


def read_inputs_rows(wb: Workbook, cfg: RunConfig) -> List[Tuple[int, str, List[str]]]:
    if "inputs" not in wb.sheetnames:
        raise ValueError("XLSX must contain an 'inputs' sheet.")

    ws = wb["inputs"]
    hm = header_map(ws)

    # required base columns
    for req in ["item_id", "source"]:
        if req not in hm:
            raise ValueError(f"'inputs' missing required column '{req}'.")

    # ensure t1..tN exist; if not, create them (blank)
    for k in range(1, cfg.num_translations + 1):
        ensure_column(ws, f"t{k}")

    # ensure row_input_hash exists (we will write it)
    ensure_column(ws, "row_input_hash")

    hm = header_map(ws)

    rows: List[Tuple[int, str, List[str]]] = []
    for r in range(2, ws.max_row + 1):
        item = ws.cell(row=r, column=hm["item_id"]).value
        if item is None or str(item).strip() == "":
            continue
        try:
            item_id = int(item)
        except Exception:
            raise ValueError(f"Non-integer item_id at row {r}: {item}")

        source = str(ws.cell(row=r, column=hm["source"]).value or "")

        translations: List[str] = []
        for k in range(1, cfg.num_translations + 1):
            translations.append(str(ws.cell(row=r, column=hm[f"t{k}"]).value or ""))

        rows.append((item_id, source, translations))

    if not rows:
        raise ValueError("No data rows found in 'inputs' sheet.")

    # sort by item_id (canonical)
    rows.sort(key=lambda x: x[0])
    return rows


def write_inputs_hashes(wb: Workbook, cfg: RunConfig, rows: List[Tuple[int, str, List[str]]]) -> None:
    ws = wb["inputs"]
    hm = header_map(ws)
    col_hash = hm["row_input_hash"]

    # Build a lookup row index by item_id
    row_by_item: Dict[int, int] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=hm["item_id"]).value
        if v is None or str(v).strip() == "":
            continue
        try:
            row_by_item[int(v)] = r
        except Exception:
            continue

    for item_id, source, translations in rows:
        h = compute_row_input_hash(source, translations)
        rr = row_by_item.get(item_id)
        if rr is None:
            continue
        ws.cell(row=rr, column=col_hash).value = h


def rebuild_eval_sheet(wb: Workbook, cfg: RunConfig, rows: List[Tuple[int, str, List[str]]], run_id: str) -> None:
    if "eval" in wb.sheetnames:
        wb.remove(wb["eval"])
    ws = wb.create_sheet("eval")

    ws.append(_eval_header(cfg))
    hm = header_map(ws)

    for (item_id, source, translations) in rows:
        translations = (translations + [""] * cfg.num_translations)[: cfg.num_translations]
        row_input_hash = compute_row_input_hash(source, translations)
        ws.append(
            [item_id]
            + [""] * cfg.num_translations
            + [""] * cfg.num_translations
            + ["", "", "", 0, "", row_input_hash, "", run_id]
        )

    ws.freeze_panes = "A2"


def repair_eval_sheet_in_place(wb: Workbook, cfg: RunConfig, rows: List[Tuple[int, str, List[str]]], run_id: str) -> None:
    """
    Keeps existing eval data if present; ensures required columns; aligns rows to inputs.item_id; fixes row_input_hash/run_id.
    """
    if "eval" not in wb.sheetnames:
        rebuild_eval_sheet(wb, cfg, rows, run_id)
        return

    ws = wb["eval"]

    # Ensure all required columns exist (append missing)
    for col in _eval_header(cfg):
        ensure_column(ws, col)

    hm = header_map(ws)

    # Map existing eval rows by item_id
    eval_row_by_item: Dict[int, int] = {}
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=hm["item_id"]).value
        if v is None or str(v).strip() == "":
            continue
        try:
            eval_row_by_item[int(v)] = r
        except Exception:
            continue

    # For each input row, ensure an eval row exists; if not, append it
    for (item_id, source, translations) in rows:
        translations = (translations + [""] * cfg.num_translations)[: cfg.num_translations]
        row_input_hash = compute_row_input_hash(source, translations)

        rr = eval_row_by_item.get(item_id)
        if rr is None:
            # append a new row with defaults
            ws.append(
                [item_id]
                + [""] * cfg.num_translations
                + [""] * cfg.num_translations
                + ["", "", "", 0, "", row_input_hash, "", run_id]
            )
            rr = ws.max_row
            eval_row_by_item[item_id] = rr

        # Repair/ensure required fields
        ws.cell(row=rr, column=hm["row_input_hash"]).value = row_input_hash
        ws.cell(row=rr, column=hm["run_id"]).value = run_id

        # Keep existing started/committed/comment/buckets/das if present; just ensure defaults when blank
        if ws.cell(row=rr, column=hm["edit_count"]).value in (None, ""):
            ws.cell(row=rr, column=hm["edit_count"]).value = 0

        # display_map_json: keep if valid-ish, else blank (main app will generate deterministically)
        dm = ws.cell(row=rr, column=hm["display_map_json"]).value
        if dm:
            try:
                m = json.loads(str(dm))
                if not isinstance(m, dict):
                    ws.cell(row=rr, column=hm["display_map_json"]).value = ""
            except Exception:
                ws.cell(row=rr, column=hm["display_map_json"]).value = ""

    # Optional: you may want to warn if eval has extra item_ids not in inputs
    # (we do not delete them automatically)


# -----------------------------
# UI
# -----------------------------
st.title("Generate Evaluation XLSX")
try:
    cfg = load_config("config.yaml")
    st.session_state.cfg = cfg
except Exception as e:
    st.error(f"Config error: {e}")
    st.stop()

st.caption(f"Configured for {cfg.num_translations} translations per sentence (t1..t{cfg.num_translations}).")
with st.expander("Run configuration (frozen for this run)", expanded=False):
    st.json(
        {
            "num_translations": cfg.num_translations,
            #"da": {"min": cfg.da_min, "max": cfg.da_max, "integer_only": cfg.da_integer_only},
            "buckets": [{"key": b.key, "label": b.label} for b in cfg.buckets],
            "validation": {
                "enforce_bucket_ordering": cfg.enforce_bucket_ordering,
                "allow_empty_buckets": cfg.allow_empty_buckets,
            },
            "da_intra_bucket_options": cfg.da_intra_bucket_options,
            "bucket_colors": cfg.bucket_colors
        }
    )

tab1, tab2, tab3 = st.tabs(["Generate example (5 rows)", "Upload CSV → XLSX", "Upload XLSX → Repair/Rebuild"])

with tab1:
    st.subheader("Example file generator")
    st.write("Creates 5 example source sentences and placeholder translations per sentence.")

    n_rows = 5
    if st.button("Generate example XLSX", type="primary"):
        sources = [
            "The quick brown fox jumps over the lazy dog.",
            "I would like to book a table for two at 7 PM.",
            "Please translate this sentence as naturally as possible.",
            "The model produced multiple candidates with varying fluency.",
            "Human evaluation should be consistent and repeatable.",
        ][:n_rows]

        t_cols: Dict[int, List[str]] = {}
        for k in range(1, cfg.num_translations + 1):
            t_cols[k] = [latin_filler(seed=10_000 * k + i) for i in range(1, n_rows + 1)]

        rows = normalize_rows_to_n_translations(sources, t_cols, cfg.num_translations)
        wb = build_workbook(rows, cfg)
        data = workbook_to_bytes(wb)

        st.download_button(
            "Download example XLSX",
            data=data,
            file_name="mt_eval_example_5rows.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

with tab2:
    st.subheader("CSV → XLSX template")
    st.write(
        "Upload a CSV containing `source` and optionally `t1..t12` (or whatever N is in config). "
        "Missing `t#` columns will be filled with empty strings."
    )

    csv_up = st.file_uploader("Upload CSV", type=["csv"], accept_multiple_files=False, key="csv_uploader")

    colA, colB = st.columns([1, 1])
    with colA:
        run_id_override = st.text_input("Optional run_id (leave blank to auto-generate)", value="", key="csv_run_id")
    with colB:
        out_name = st.text_input("Output filename", value="mt_eval_from_csv.xlsx", key="csv_out_name")

    if csv_up is not None:
        try:
            rows = parse_csv(csv_up.read(), cfg)

            st.success(f"Loaded {len(rows)} row(s) from CSV.")
            st.write("Preview (first 5):")
            preview = []
            for (item_id, src, ts) in rows[:5]:
                preview.append({"item_id": item_id, "source": src, "t1": ts[0] if ts else ""})
            st.dataframe(preview, use_container_width=True)

            wb = build_workbook(rows, cfg, run_id=run_id_override.strip() or None)
            data = workbook_to_bytes(wb)

            st.download_button(
                "Download XLSX template",
                data=data,
                file_name=out_name if out_name.endswith(".xlsx") else out_name + ".xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        except Exception as e:
            st.error(f"Failed to generate XLSX: {e}")

with tab3:
    st.subheader("Upload XLSX → Repair/Rebuild")
    st.write(
        "Upload an XLSX that contains an `inputs` sheet. This tool will recompute `row_input_hash` for each row. "
        "You can also rebuild or repair the `eval` sheet so the main rating app can load it reliably."
    )

    xlsx_up = st.file_uploader("Upload XLSX", type=["xlsx"], accept_multiple_files=False, key="xlsx_uploader")

    c1, c2, c3 = st.columns([1, 1, 1.2])
    with c1:
        rebuild_eval = st.checkbox("Rebuild eval sheet from scratch", value=False)
    with c2:
        run_id_override_xlsx = st.text_input("Optional run_id override", value="", key="xlsx_run_id")
    with c3:
        out_xlsx_name = st.text_input("Output filename", value="mt_eval_repaired.xlsx", key="xlsx_out_name")

    if xlsx_up is not None:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(xlsx_up.read()))
            rows = read_inputs_rows(wb, cfg)

            # Determine run_id
            run_id = run_id_override_xlsx.strip()
            if not run_id:
                # If eval exists and has a run_id value, keep it
                run_id = str(uuid.uuid4())
                if "eval" in wb.sheetnames:
                    ev = wb["eval"]
                    hm = header_map(ev)
                    if "run_id" in hm:
                        # pick first non-empty run_id
                        for r in range(2, ev.max_row + 1):
                            v = ev.cell(row=r, column=hm["run_id"]).value
                            if v and str(v).strip():
                                run_id = str(v).strip()
                                break

            # Write/repair inputs hashes
            write_inputs_hashes(wb, cfg, rows)

            # Rebuild or repair eval
            if rebuild_eval:
                rebuild_eval_sheet(wb, cfg, rows, run_id)
            else:
                repair_eval_sheet_in_place(wb, cfg, rows, run_id)

            data = workbook_to_bytes(wb)

            st.success(
                f"Processed {len(rows)} input row(s). "
                f"{'Rebuilt' if rebuild_eval else 'Repaired'} eval sheet. "
                f"run_id={run_id}"
            )

            st.download_button(
                "Download repaired XLSX",
                data=data,
                file_name=out_xlsx_name if out_xlsx_name.endswith(".xlsx") else out_xlsx_name + ".xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            with st.expander("Preview (first 5 rows)"):
                preview = []
                for item_id, src, ts in rows[:5]:
                    preview.append(
                        {
                            "item_id": item_id,
                            "source": src,
                            "t1": ts[0] if ts else "",
                            "row_input_hash": compute_row_input_hash(src, (ts + [""] * cfg.num_translations)[: cfg.num_translations]),
                        }
                    )
                st.dataframe(preview, use_container_width=True)

        except Exception as e:
            st.error(f"Failed to repair/rebuild XLSX: {e}")
